"""
Student Data Lookup System - FINAL VERSION
-------------------------------------------
Complete system with professional Excel formatting and auto-generated filenames.

Usage:
    python student_data_lookup.py --id 1962818
    python student_data_lookup.py --name "John Doe"
    python student_data_lookup.py --id 1962818 --output custom_name.xlsx

Features:
- Auto-generates filename: FirstName_LastName_TIMESTAMP.xlsx
- Professional Excel formatting with colors and styles
- 3 sheets: Student Profile, Events Summary, Detailed Report
- Color-coded sections for easy reading
"""

import pandas as pd
import numpy as np
import re
import argparse
from pathlib import Path
from datetime import datetime


class StudentDataLoader:
    """Loads and processes all student-related data files."""
    
    def __init__(self, data_dir='.'):
        self.data_dir = Path(data_dir)
        self.students = None
        self.signoffs = None
        self.evaluations = None
        self.events = None
        self.event_objects = None
        self.staff_objects = None
        
    def parse_sql_insert_csv(self, filepath):
        """Parse SQL INSERT format CSV into pandas DataFrame."""
        print(f"Parsing SQL format file: {filepath.name}")
        
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        records = []
        pattern = r'\(([^)]+(?:\([^)]*\)[^)]*)*)\)'
        
        for match in re.finditer(pattern, content):
            record_str = match.group(1)
            values = []
            current = ''
            in_quote = False
            
            for i, char in enumerate(record_str):
                if char == "'" and (i == 0 or record_str[i-1] != '\\'):
                    in_quote = not in_quote
                    current += char
                elif char == ',' and not in_quote:
                    values.append(current.strip())
                    current = ''
                else:
                    current += char
            
            if current:
                values.append(current.strip())
            
            clean_values = []
            for v in values:
                if v == 'NULL':
                    clean_values.append(None)
                elif v.startswith("'") and v.endswith("'"):
                    clean_values.append(v[1:-1])
                else:
                    clean_values.append(v)
            
            records.append(clean_values)
        
        df = pd.DataFrame(records, columns=[
            'id', 'account_id', 'person_id', 'first_name', 'last_name', 
            'email', 'status', 'phone', 'created_date', 'code', 'field_10',
            'sequence', 'flag', 'field_13', 'preferred_name', 'field_15',
            'field_16', 'field_17'
        ])
        
        df['id'] = pd.to_numeric(df['id'], errors='coerce')
        df['person_id'] = pd.to_numeric(df['person_id'], errors='coerce')
        df['sequence'] = pd.to_numeric(df['sequence'], errors='coerce')
        
        print(f"  ✓ Loaded {len(df)} student records")
        return df
    
    def load_students(self):
        """Load students.csv file."""
        filepath = self.data_dir / 'students.csv'
        self.students = self.parse_sql_insert_csv(filepath)
        self.students['full_name'] = (
            self.students['first_name'] + ' ' + self.students['last_name']
        )
        return self.students
    
    def load_signoffs(self):
        """Load All Campus Signoffs 2025.csv."""
        filepath = self.data_dir / 'All Campus Signoffs 2025.csv'
        print(f"Loading: {filepath.name}")
        
        self.signoffs = pd.read_csv(filepath)
        self.signoffs.columns = self.signoffs.columns.str.strip()
        
        if 'date' in self.signoffs.columns:
            self.signoffs['date'] = pd.to_datetime(self.signoffs['date'], errors='coerce')
        
        print(f"  ✓ Loaded {len(self.signoffs)} signoff records")
        return self.signoffs
    
    def load_evaluations(self):
        """Load Evaluations All Students (CSV or Excel)."""
        csv_path = self.data_dir / 'Evaluations All Students.csv'
        xlsx_path = self.data_dir / 'Evaluations All Students.xlsx'
        
        if csv_path.exists():
            print(f"Loading: {csv_path.name}")
            self.evaluations = pd.read_csv(csv_path, low_memory=False)
            
            if 'create date' in self.evaluations.columns:
                self.evaluations['create_date'] = pd.to_datetime(
                    self.evaluations['create date'], errors='coerce'
                )
        elif xlsx_path.exists():
            print(f"Loading: {xlsx_path.name}")
            self.evaluations = pd.read_excel(xlsx_path, sheet_name='Append1')
            
            if 'create date' in self.evaluations.columns:
                if pd.api.types.is_numeric_dtype(self.evaluations['create date']):
                    self.evaluations['create_date'] = pd.to_datetime(
                        '1899-12-30'
                    ) + pd.to_timedelta(self.evaluations['create date'], 'D')
                elif pd.api.types.is_datetime64_any_dtype(self.evaluations['create date']):
                    self.evaluations['create_date'] = self.evaluations['create date']
                else:
                    self.evaluations['create_date'] = pd.to_datetime(
                        self.evaluations['create date'], errors='coerce'
                    )
        else:
            raise FileNotFoundError("Could not find Evaluations file (CSV or XLSX)")
        
        print(f"  ✓ Loaded {len(self.evaluations)} evaluation records")
        return self.evaluations
    
    def load_events(self):
        """Load All Campus Forest Events (CSV or Excel)."""
        csv_path = self.data_dir / 'All Campus Forest Events 2025.csv'
        xlsx_path = self.data_dir / 'All Campus Forest Events 2025.xlsx'
        
        if csv_path.exists():
            print(f"Loading: {csv_path.name}")
            self.events = pd.read_csv(csv_path, low_memory=False)
            
            if 'start' in self.events.columns:
                self.events['start'] = pd.to_datetime(self.events['start'], errors='coerce')
            if 'end' in self.events.columns:
                self.events['end'] = pd.to_datetime(self.events['end'], errors='coerce')
            
            campuses = self.events['campus'].nunique() if 'campus' in self.events.columns else 'unknown'
            print(f"  ✓ Loaded {len(self.events)} event records from {campuses} campuses")
        elif xlsx_path.exists():
            print(f"Loading: {xlsx_path.name}")
            excel_file = pd.ExcelFile(xlsx_path)
            all_events = []
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
                df['campus'] = sheet_name
                all_events.append(df)
            
            self.events = pd.concat(all_events, ignore_index=True)
            
            if 'start' in self.events.columns:
                self.events['start'] = pd.to_datetime(self.events['start'], errors='coerce')
            if 'end' in self.events.columns:
                self.events['end'] = pd.to_datetime(self.events['end'], errors='coerce')
            
            print(f"  ✓ Loaded {len(self.events)} event records from {len(excel_file.sheet_names)} campuses")
        else:
            raise FileNotFoundError("Could not find Events file (CSV or XLSX)")
        
        return self.events
    
    def load_event_objects(self):
        """Load Event Objects/Vehicles (CSV or Excel)."""
        possible_files = [
            ('Event Objects.csv', 'csv'),
            ('Event Objects.xlsx', 'xlsx'),
            ('2025 Event ObjectsVehicles.csv', 'csv'),
            ('2025 Event ObjectsVehicles.xlsx', 'xlsx'),
            ('Event ObjectsVehicles.csv', 'csv'),
            ('Event ObjectsVehicles.xlsx', 'xlsx'),
        ]
        
        for filename, file_type in possible_files:
            filepath = self.data_dir / filename
            if filepath.exists():
                print(f"Loading: {filepath.name}")
                if file_type == 'csv':
                    self.event_objects = pd.read_csv(filepath)
                else:
                    self.event_objects = pd.read_excel(filepath)
                print(f"  ✓ Loaded {len(self.event_objects)} event object records")
                return self.event_objects
        
        raise FileNotFoundError("Could not find Event Objects file")
    
    def load_staff_objects(self):
        """Load Staff Objects (CSV or Excel)."""
        possible_files = [
            ('Staff Objects.csv', 'csv'),
            ('Staff Objects.xlsx', 'xlsx'),
            ('2025 Staff Objects.csv', 'csv'),
            ('2025 Staff Objects.xlsx', 'xlsx'),
        ]
        
        for filename, file_type in possible_files:
            filepath = self.data_dir / filename
            if filepath.exists():
                print(f"Loading: {filepath.name}")
                if file_type == 'csv':
                    self.staff_objects = pd.read_csv(filepath)
                else:
                    self.staff_objects = pd.read_excel(filepath)
                print(f"  ✓ Loaded {len(self.staff_objects)} staff object records")
                return self.staff_objects
        
        print("  ⚠️  Warning: Staff Objects file not found - trainer names will not be available")
        self.staff_objects = pd.DataFrame()
        return self.staff_objects
    
    def load_all(self):
        """Load all data files."""
        print("\n" + "="*60)
        print("LOADING ALL DATA FILES")
        print("="*60 + "\n")
        
        self.load_students()
        self.load_signoffs()
        self.load_evaluations()
        self.load_events()
        self.load_event_objects()
        self.load_staff_objects()
        
        print("\n" + "="*60)
        print("ALL FILES LOADED SUCCESSFULLY")
        print("="*60 + "\n")


class StudentProfiler:
    """Creates comprehensive student profiles from loaded data."""
    
    def __init__(self, loader):
        self.loader = loader
    
    def get_student_by_id(self, student_id):
        """Get student record by ID."""
        student = self.loader.students[self.loader.students['id'] == student_id]
        return student.iloc[0] if len(student) > 0 else None
    
    def get_student_by_name(self, name):
        """Get student record by name (partial match)."""
        name_lower = name.lower()
        matches = self.loader.students[
            self.loader.students['full_name'].str.lower().str.contains(name_lower, na=False)
        ]
        return matches
    
    def get_signoffs(self, student_name):
        """Get all signoffs for a student by name."""
        signoffs = self.loader.signoffs[
            self.loader.signoffs['submitter'].str.contains(student_name, case=False, na=False)
        ]
        return signoffs
    
    def get_evaluations(self, student_id):
        """Get all evaluations for a student by ID."""
        evals = self.loader.evaluations[self.loader.evaluations['Student'] == student_id]
        return evals
    
    def get_events_for_student(self, student_id):
        """Get all events associated with a student through evaluations."""
        evals = self.get_evaluations(student_id)
        if len(evals) == 0:
            return pd.DataFrame()
        
        event_ids = evals['event'].unique()
        events = self.loader.events[self.loader.events['event id'].isin(event_ids)]
        return events
    
    def get_event_objects_for_events(self, event_ids):
        """Get all objects/vehicles for given event IDs."""
        if len(event_ids) == 0:
            return pd.DataFrame()
        
        objects = self.loader.event_objects[self.loader.event_objects['event'].isin(event_ids)]
        return objects
    
    def build_complete_profile(self, student_id=None, student_name=None):
        """Build a comprehensive profile for a student."""
        if student_id:
            student_record = self.get_student_by_id(student_id)
            if student_record is None:
                return None
            search_name = f"{student_record['first_name']} {student_record['last_name']}"
        elif student_name:
            matches = self.get_student_by_name(student_name)
            if len(matches) == 0:
                return None
            elif len(matches) > 1:
                print(f"\nFound {len(matches)} students matching '{student_name}':")
                for idx, row in matches.iterrows():
                    print(f"  - ID: {row['id']}, Name: {row['full_name']}, Status: {row['status']}")
                return matches
            student_record = matches.iloc[0]
            student_id = student_record['id']
            search_name = student_record['full_name']
        else:
            return None
        
        profile = {
            'student': student_record,
            'signoffs': self.get_signoffs(search_name),
            'evaluations': self.get_evaluations(student_id),
            'events': self.get_events_for_student(student_id)
        }
        
        if len(profile['events']) > 0:
            event_ids = profile['events']['event id'].unique()
            profile['event_objects'] = self.get_event_objects_for_events(event_ids)
        else:
            profile['event_objects'] = pd.DataFrame()
        
        return profile
    
    def print_profile_summary(self, profile):
        """Print a summary of the student profile."""
        if profile is None:
            print("No student found.")
            return
        
        if isinstance(profile, pd.DataFrame):
            print(f"Multiple students found. Please use specific student ID.")
            return
        
        student = profile['student']
        
        print("\n" + "="*60)
        print(f"STUDENT PROFILE: {student['full_name']}")
        print("="*60)
        print(f"ID: {student['id']}")
        print(f"Email: {student['email']}")
        print(f"Phone: {student['phone']}")
        print(f"Status: {student['status']}")
        print(f"Code: {student['code']}")
        print(f"Created: {student['created_date']}")
        if student['preferred_name']:
            print(f"Preferred Name: {student['preferred_name']}")
        
        print("\n" + "-"*60)
        print("SUMMARY STATISTICS")
        print("-"*60)
        print(f"Total Signoffs: {len(profile['signoffs'])}")
        print(f"Total Evaluations: {len(profile['evaluations'])}")
        print(f"Total Events: {len(profile['events'])}")
        print(f"Total Event Objects: {len(profile['event_objects'])}")
        
        if len(profile['evaluations']) > 0:
            avg_score = profile['evaluations']['score'].mean()
            print(f"\nAverage Evaluation Score: {avg_score:.2f}")
            
            skill_counts = profile['evaluations']['skill group'].value_counts().head(5)
            print("\nTop 5 Evaluated Skills:")
            for skill, count in skill_counts.items():
                avg_skill_score = profile['evaluations'][
                    profile['evaluations']['skill group'] == skill
                ]['score'].mean()
                print(f"  - {skill}: {count} evaluations (avg: {avg_skill_score:.2f})")
        
        if len(profile['signoffs']) > 0:
            agree_count = (profile['signoffs']['state'] == 'agree').sum()
            disagree_count = (profile['signoffs']['state'] == 'disagree').sum()
            print(f"\nSignoff States:")
            print(f"  - Agree: {agree_count}")
            print(f"  - Disagree: {disagree_count}")
        
        print("\n" + "="*60 + "\n")
    
    def remove_timezone_from_dataframe(self, df):
        """Remove timezone information from all datetime columns."""
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                if hasattr(df_copy[col].dtype, 'tz') and df_copy[col].dtype.tz is not None:
                    df_copy[col] = df_copy[col].dt.tz_localize(None)
        return df_copy
    
    def build_event_centric_report(self, profile):
        """Build a report organized by events with all related data."""
        if profile is None or isinstance(profile, pd.DataFrame):
            return None
        
        events_data = []
        events = profile['events'].copy()
        
        for idx, event_row in events.iterrows():
            event_id = event_row.get('event id')
            
            # Get trainer name from Staff Objects
            trainer_name = ''
            if self.loader.staff_objects is not None and len(self.loader.staff_objects) > 0:
                staff_for_event = self.loader.staff_objects[
                    (self.loader.staff_objects['event'] == event_id) &
                    (self.loader.staff_objects['state'] == 'active')
                ]
                if len(staff_for_event) > 0:
                    # Get all staff names for this event, separated by commas
                    trainer_names = staff_for_event['object'].tolist()
                    trainer_name = ', '.join([str(name) for name in trainer_names if pd.notna(name)])
            
            event_info = {
                'Event ID': event_id,
                'Office': event_row.get('office', ''),
                'Module': event_row.get('module', event_row.get('staff title', '')),
                'Location': event_row.get('location', event_row.get('campus', '')),
                'Event Date/Time': event_row.get('event', ''),
                'Odometer Start': event_row.get('odometer start', ''),
                'Odometer End': event_row.get('odometer end', ''),
                'Progress State': event_row.get('progress state', ''),
                'State': event_row.get('state', ''),
                'Staff Sign Off': trainer_name,
                'Campus': event_row.get('campus', ''),
            }
            
            event_evals = profile['evaluations'][
                profile['evaluations']['event'] == event_id
            ].copy()
            
            if len(event_evals) > 0:
                event_info['Evaluation Count'] = len(event_evals)
                event_info['Average Score'] = event_evals['score'].mean()
                
                skills_data = []
                for _, eval_row in event_evals.iterrows():
                    skills_data.append({
                        'Skill Group': eval_row.get('skill group', ''),
                        'Score': eval_row.get('score', ''),
                        'Notes': eval_row.get('notes', eval_row.get('comment', '')),
                    })
                event_info['Skills'] = skills_data
            else:
                event_info['Evaluation Count'] = 0
                event_info['Average Score'] = ''
                event_info['Skills'] = []
            
            event_objects = profile['event_objects'][
                profile['event_objects']['event'] == event_id
            ].copy()
            
            objects_data = []
            for _, obj_row in event_objects.iterrows():
                objects_data.append({
                    'Object': obj_row.get('object', ''),
                    'Type': obj_row.get('type', ''),
                    'State': obj_row.get('state', ''),
                })
            event_info['Objects'] = objects_data
            event_info['Object Count'] = len(objects_data)
            
            event_start = pd.to_datetime(event_row.get('start'), errors='coerce')
            if pd.notna(event_start):
                event_date = event_start.date()
                signoffs_for_event = profile['signoffs'][
                    pd.to_datetime(profile['signoffs']['date'], errors='coerce').dt.date == event_date
                ].copy()
            else:
                signoffs_for_event = pd.DataFrame()
            
            signoffs_data = []
            for _, sign_row in signoffs_for_event.iterrows():
                signoffs_data.append({
                    'Signature': sign_row.get('signature', ''),
                    'Signoff ID': sign_row.get('signoff id', ''),
                    'State': sign_row.get('state', ''),
                    'Submitter': sign_row.get('submitter', ''),
                })
            event_info['Signoffs'] = signoffs_data
            event_info['Signoff Count'] = len(signoffs_data)
            
            events_data.append(event_info)
        
        return events_data
    
    def export_event_centric_to_excel(self, profile, output_file):
        """Export event-centric report to Excel with professional formatting."""
        if profile is None or isinstance(profile, pd.DataFrame):
            print("Cannot export: Invalid profile data")
            return
        
        events_data = self.build_event_centric_report(profile)
        if not events_data:
            print("No events found for this student")
            return
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import ColorScaleRule
            print("  🎨 Applying professional formatting...")
            formatting_available = True
        except ImportError:
            print("  ⚠️  Formatting libraries not available, creating basic Excel...")
            formatting_available = False
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            student = profile['student']
            
            # SHEET 1: Student Profile Cover Page
            info_data = {
                'Field': ['Student ID', 'Full Name', 'Preferred Name', 'Email', 'Phone', 'Status', 
                          'Code', 'Created Date', '', 
                          'Total Events', 'Total Evaluations', 'Total Signoffs', 'Total Objects Used',
                          'Average Score', 'Agree Signoffs', 'Disagree Signoffs'],
                'Value': [
                    student['id'],
                    f"{student['first_name']} {student['last_name']}",
                    student.get('preferred_name', 'N/A'),
                    student.get('email', 'N/A'),
                    student.get('phone', 'N/A'),
                    student.get('status', 'N/A'),
                    student.get('code', 'N/A'),
                    student.get('created_date', 'N/A'),
                    '',
                    len(profile['events']),
                    len(profile['evaluations']),
                    len(profile['signoffs']),
                    len(profile['event_objects']),
                    f"{profile['evaluations']['score'].mean():.2f}" if len(profile['evaluations']) > 0 else 'N/A',
                    (profile['signoffs']['state'] == 'agree').sum() if len(profile['signoffs']) > 0 else 0,
                    (profile['signoffs']['state'] == 'disagree').sum() if len(profile['signoffs']) > 0 else 0,
                ]
            }
            
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='👤 Student Profile', index=False)
            
            if formatting_available:
                ws_info = writer.sheets['👤 Student Profile']
                ws_info.insert_rows(1)
                ws_info.merge_cells('A1:B1')
                ws_info['A1'] = f"STUDENT PROFILE: {student['first_name']} {student['last_name']}"
                ws_info['A1'].font = Font(size=16, bold=True, color='FFFFFF')
                ws_info['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                ws_info['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws_info.row_dimensions[1].height = 35
                
                ws_info['A2'].font = Font(bold=True, size=11, color='FFFFFF')
                ws_info['A2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                ws_info['B2'].font = Font(bold=True, size=11, color='FFFFFF')
                ws_info['B2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                
                info_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                stats_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                
                for row_idx in range(3, ws_info.max_row + 1):
                    field_cell = ws_info[f'A{row_idx}']
                    value_cell = ws_info[f'B{row_idx}']
                    field_cell.font = Font(bold=True, size=10)
                    field_cell.alignment = Alignment(horizontal='right', vertical='center')
                    value_cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    if row_idx >= 12:
                        field_cell.fill = stats_fill
                        value_cell.fill = stats_fill
                    else:
                        field_cell.fill = info_fill
                        value_cell.fill = info_fill
                
                ws_info.column_dimensions['A'].width = 25
                ws_info.column_dimensions['B'].width = 40
            
            # SHEET 2: Events Summary
            summary_rows = []
            for event_info in events_data:
                summary_rows.append({
                    'Event ID': event_info['Event ID'],
                    'Office': event_info['Office'],
                    'Module': event_info['Module'],
                    'Location': event_info['Location'],
                    'Event Date/Time': event_info['Event Date/Time'],
                    'Progress State': event_info['Progress State'],
                    'Evaluation Count': event_info['Evaluation Count'],
                    'Average Score': event_info['Average Score'],
                    'Object Count': event_info['Object Count'],
                    'Signoff Count': event_info['Signoff Count'],
                })
            
            summary_df = pd.DataFrame(summary_rows)
            summary_df = self.remove_timezone_from_dataframe(summary_df)
            
            for col in summary_df.columns:
                if summary_df[col].dtype == 'object':
                    summary_df[col] = summary_df[col].apply(
                        lambda x: str(x) if pd.notna(x) and hasattr(x, 'tzinfo') else x
                    )
            
            summary_df.to_excel(writer, sheet_name='📊 Events Summary', index=False)
            
            if formatting_available:
                ws_summary = writer.sheets['📊 Events Summary']
                header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)
                
                for cell in ws_summary[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for column in ws_summary.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_summary.column_dimensions[column_letter].width = adjusted_width
                
                ws_summary.freeze_panes = 'A2'
                ws_summary.auto_filter.ref = ws_summary.dimensions
                
                score_col = None
                for idx, cell in enumerate(ws_summary[1], 1):
                    if cell.value == 'Average Score':
                        score_col = get_column_letter(idx)
                        break
                
                if score_col:
                    ws_summary.conditional_formatting.add(
                        f'{score_col}2:{score_col}{len(summary_df)+1}',
                        ColorScaleRule(
                            start_type='num', start_value=0, start_color='F8696B',
                            mid_type='num', mid_value=3, mid_color='FFEB84',
                            end_type='num', end_value=5, end_color='63BE7B'
                        )
                    )
            
            # SHEET 3: Detailed Report (Compact View)
            compact_rows = []
            for event_info in events_data:
                compact_rows.append({
                    'Event ID': event_info['Event ID'],
                    'Office': event_info['Office'],
                    'Module': event_info['Module'],
                    'Location': event_info['Location'],
                    'Event Date/Time': event_info['Event Date/Time'],
                    'Odometer Start': event_info['Odometer Start'],
                    'Odometer End': event_info['Odometer End'],
                    'Progress State': event_info['Progress State'],
                    'Staff Sign Off': event_info['Staff Sign Off'],
                    'Skill Group': '',
                    'Score': '',
                    'Notes': '',
                    'Object': '',
                    'Object Type': '',
                    'Signature': '',
                    'Signoff ID': '',
                    'Signoff State': '',
                    'Submitter': '',
                })
                
                max_rows = max(
                    len(event_info['Skills']),
                    len(event_info['Objects']),
                    len(event_info['Signoffs']),
                    1
                )
                
                for i in range(max_rows):
                    row = {
                        'Event ID': '' if i > 0 else event_info['Event ID'],
                        'Office': '',
                        'Module': '',
                        'Location': '',
                        'Event Date/Time': '',
                        'Odometer Start': '',
                        'Odometer End': '',
                        'Progress State': '',
                        'Staff Sign Off': '',
                    }
                    
                    if i < len(event_info['Skills']):
                        skill = event_info['Skills'][i]
                        row['Skill Group'] = skill['Skill Group']
                        row['Score'] = skill['Score']
                        row['Notes'] = skill['Notes']
                    else:
                        row['Skill Group'] = ''
                        row['Score'] = ''
                        row['Notes'] = ''
                    
                    if i < len(event_info['Objects']):
                        obj = event_info['Objects'][i]
                        row['Object'] = obj['Object']
                        row['Object Type'] = obj['Type']
                    else:
                        row['Object'] = ''
                        row['Object Type'] = ''
                    
                    if i < len(event_info['Signoffs']):
                        signoff = event_info['Signoffs'][i]
                        row['Signature'] = signoff['Signature']
                        row['Signoff ID'] = signoff['Signoff ID']
                        row['Signoff State'] = signoff['State']
                        row['Submitter'] = signoff['Submitter']
                    else:
                        row['Signature'] = ''
                        row['Signoff ID'] = ''
                        row['Signoff State'] = ''
                        row['Submitter'] = ''
                    
                    if i > 0:
                        compact_rows.append(row)
                
                compact_rows.append({})
            
            compact_df = pd.DataFrame(compact_rows)
            compact_df = self.remove_timezone_from_dataframe(compact_df)
            
            for col in compact_df.columns:
                if compact_df[col].dtype == 'object':
                    compact_df[col] = compact_df[col].apply(
                        lambda x: str(x) if pd.notna(x) and hasattr(x, 'tzinfo') else x
                    )
            
            compact_df.to_excel(writer, sheet_name='📋 Detailed Report', index=False)
            
            if formatting_available:
                ws_compact = writer.sheets['📋 Detailed Report']
                
                event_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                eval_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                object_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                signoff_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                
                for cell in ws_compact[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
                
                event_font = Font(bold=True, size=10)
                
                for row_idx, row in enumerate(ws_compact.iter_rows(min_row=2, max_row=ws_compact.max_row), start=2):
                    if row[0].value and str(row[0].value).strip():
                        for cell in row[:9]:
                            cell.fill = event_fill
                            cell.font = event_font
                            cell.border = thin_border
                            cell.alignment = Alignment(vertical='center')
                    else:
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = Alignment(vertical='center')
                        
                        if row[9].value:
                            for cell in row[9:12]:
                                cell.fill = eval_fill
                        
                        if row[12].value:
                            for cell in row[12:14]:
                                cell.fill = object_fill
                        
                        if row[14].value:
                            for cell in row[14:18]:
                                cell.fill = signoff_fill
                
                for column in ws_compact.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws_compact.column_dimensions[column_letter].width = adjusted_width
                
                ws_compact.freeze_panes = 'A2'
                ws_compact.auto_filter.ref = ws_compact.dimensions
                ws_compact.row_dimensions[1].height = 30
        
        print(f"✓ Event-centric report exported to: {output_file}")
        if formatting_available:
            print(f"  📊 3 sheets created:")
            print(f"     1. 👤 Student Profile - Overview with statistics")
            print(f"     2. 📊 Events Summary - Quick overview with score heatmap")  
            print(f"     3. 📋 Detailed Report - Complete event details (MAIN SHEET)")
            print(f"  🎨 Professional formatting:")
            print(f"     • Color-coded: Events (blue), Evaluations (green), Objects (orange), Signoffs (yellow)")
            print(f"     • Auto-sized columns, frozen headers, sortable filters")


def main():
    parser = argparse.ArgumentParser(description='Student Data Lookup System')
    parser.add_argument('--name', type=str, help='Search by student name')
    parser.add_argument('--id', type=int, help='Search by student ID')
    parser.add_argument('--batch', action='store_true', help='Process all students')
    parser.add_argument('--output', type=str, default=None, 
                        help='Output Excel file name (default: auto-generated with timestamp)')
    parser.add_argument('--dir', type=str, default='.', 
                        help='Directory containing data files')
    
    args = parser.parse_args()
    
    loader = StudentDataLoader(data_dir=args.dir)
    loader.load_all()
    
    profiler = StudentProfiler(loader)
    
    if args.batch:
        print("Batch processing not yet implemented. Use --name or --id for individual queries.")
    elif args.name:
        profile = profiler.build_complete_profile(student_name=args.name)
        if profile is not None and not isinstance(profile, pd.DataFrame):
            profiler.print_profile_summary(profile)
            
            if args.output is None:
                student = profile['student']
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                first = str(student['first_name']).replace(' ', '_').replace('/', '_')
                last = str(student['last_name']).replace(' ', '_').replace('/', '_')
                args.output = f"{first}_{last}_{timestamp}.xlsx"
                print(f"\n📁 Creating report: {args.output}")
            
            profiler.export_event_centric_to_excel(profile, args.output)
        elif isinstance(profile, pd.DataFrame):
            print(f"\nFound {len(profile)} students. Please refine search or use --id")
    elif args.id:
        profile = profiler.build_complete_profile(student_id=args.id)
        if profile is not None:
            profiler.print_profile_summary(profile)
            
            if args.output is None:
                student = profile['student']
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                first = str(student['first_name']).replace(' ', '_').replace('/', '_')
                last = str(student['last_name']).replace(' ', '_').replace('/', '_')
                args.output = f"{first}_{last}_{timestamp}.xlsx"
                print(f"\n📁 Creating report: {args.output}")
            
            profiler.export_event_centric_to_excel(profile, args.output)
        else:
            print(f"No student found with ID: {args.id}")
    else:
        print("Please specify --name or --id to search for a student.")
        print("Example: python student_data_lookup.py --name 'John Doe'")
        print("Example: python student_data_lookup.py --id 1694071")


if __name__ == "__main__":
    main()
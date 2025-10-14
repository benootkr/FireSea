"""
Student Data Lookup Web Application
------------------------------------
Beautiful web interface for generating student reports.

To run: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from pathlib import Path
from datetime import datetime
import io

# Page configuration
st.set_page_config(
    page_title="Student Data Lookup",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        color: #1F4E78;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
    }
    .stat-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1F4E78;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #28a745;
    }
    .stButton>button {
        width: 100%;
        background-color: #1F4E78;
        color: white;
        font-size: 1.1rem;
        padding: 0.5rem;
        border-radius: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)


class StudentDataLoader:
    """Loads and processes all student-related data files."""
    
    def __init__(self, data_dir='.'):
        self.data_dir = Path(data_dir)
        self.students = None
        self.signoffs = None
        self.evaluations = None
        self.events = None
        self.event_objects = None
        self.staff_objects = None
        self.invoices = None
        self.payments = None
        self.enrollments = None
        
    def parse_sql_insert_csv(self, filepath):
        """Parse SQL INSERT format CSV into pandas DataFrame."""
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        records = []
        pattern = r'\(([^)]+(?:\([^)]*\)[^)]*)*)\)'
        
        for match in re.finditer(pattern, content):
            record_str = match.group(1)
            values = []
            current = ''
            in_quote = False
            
            for i, char in enumerate(record_str):
                if char == "'" and (i == 0 or record_str[i-1] != '\\'):
                    in_quote = not in_quote
                    current += char
                elif char == ',' and not in_quote:
                    values.append(current.strip())
                    current = ''
                else:
                    current += char
            
            if current:
                values.append(current.strip())
            
            clean_values = []
            for v in values:
                if v == 'NULL':
                    clean_values.append(None)
                elif v.startswith("'") and v.endswith("'"):
                    clean_values.append(v[1:-1])
                else:
                    clean_values.append(v)
            
            records.append(clean_values)
        
        df = pd.DataFrame(records, columns=[
            'id', 'account_id', 'person_id', 'first_name', 'last_name', 
            'email', 'status', 'phone', 'created_date', 'code', 'field_10',
            'sequence', 'flag', 'field_13', 'preferred_name', 'field_15',
            'field_16', 'field_17'
        ])
        
        df['id'] = pd.to_numeric(df['id'], errors='coerce')
        df['person_id'] = pd.to_numeric(df['person_id'], errors='coerce')
        df['sequence'] = pd.to_numeric(df['sequence'], errors='coerce')
        
        return df
    
    def load_students(self):
        """Load students.csv file."""
        filepath = self.data_dir / 'students.csv'
        self.students = self.parse_sql_insert_csv(filepath)
        self.students['full_name'] = (
            self.students['first_name'] + ' ' + self.students['last_name']
        )
        return self.students
    
    def load_signoffs(self):
        """Load All Campus Signoffs 2025.csv."""
        filepath = self.data_dir / 'All Campus Signoffs 2025.csv'
        self.signoffs = pd.read_csv(filepath)
        self.signoffs.columns = self.signoffs.columns.str.strip()
        
        if 'date' in self.signoffs.columns:
            self.signoffs['date'] = pd.to_datetime(self.signoffs['date'], errors='coerce')
        
        return self.signoffs
    
    def load_evaluations(self):
        """Load Evaluations All Students (CSV or Excel)."""
        csv_path = self.data_dir / 'Evaluations All Students.csv'
        xlsx_path = self.data_dir / 'Evaluations All Students.xlsx'
        
        if csv_path.exists():
            self.evaluations = pd.read_csv(csv_path, low_memory=False)
            
            if 'create date' in self.evaluations.columns:
                self.evaluations['create_date'] = pd.to_datetime(
                    self.evaluations['create date'], errors='coerce'
                )
        elif xlsx_path.exists():
            self.evaluations = pd.read_excel(xlsx_path, sheet_name='Append1')
            
            if 'create date' in self.evaluations.columns:
                if pd.api.types.is_numeric_dtype(self.evaluations['create date']):
                    self.evaluations['create_date'] = pd.to_datetime(
                        '1899-12-30'
                    ) + pd.to_timedelta(self.evaluations['create date'], 'D')
                elif pd.api.types.is_datetime64_any_dtype(self.evaluations['create date']):
                    self.evaluations['create_date'] = self.evaluations['create date']
                else:
                    self.evaluations['create_date'] = pd.to_datetime(
                        self.evaluations['create date'], errors='coerce'
                    )
        
        return self.evaluations
    
    def load_events(self):
        """Load All Campus Forest Events (CSV or Excel)."""
        csv_path = self.data_dir / 'All Campus Forest Events 2025.csv'
        
        if csv_path.exists():
            self.events = pd.read_csv(csv_path, low_memory=False)
            
            if 'start' in self.events.columns:
                self.events['start'] = pd.to_datetime(self.events['start'], errors='coerce')
            if 'end' in self.events.columns:
                self.events['end'] = pd.to_datetime(self.events['end'], errors='coerce')
        
        return self.events
    
    def load_event_objects(self):
        """Load Event Objects."""
        filepath = self.data_dir / '2025 Event ObjectsVehicles.csv'
        if filepath.exists():
            self.event_objects = pd.read_csv(filepath)
            
            # Debug: Print column names to see what we have
            print(f"Event Objects columns: {self.event_objects.columns.tolist()}")
            
            # Clean column names (remove extra spaces)
            self.event_objects.columns = self.event_objects.columns.str.strip()
            
            # ===== FIX: Ensure 'event' column is numeric for proper matching =====
            if 'event' in self.event_objects.columns:
                self.event_objects['event'] = pd.to_numeric(self.event_objects['event'], errors='coerce')
                print(f"Loaded {len(self.event_objects)} event objects")
            else:
                print(f"WARNING: 'event' column not found in Event Objects CSV!")
                print(f"Available columns: {self.event_objects.columns.tolist()}")
            # ===== END FIX =====
        else:
            print(f"Event Objects file not found: {filepath}")
            self.event_objects = pd.DataFrame()
        return self.event_objects
    
    def load_staff_objects(self):
        """Load Staff Objects."""
        filepath = self.data_dir / 'Staff Objects.csv'
        if filepath.exists():
            self.staff_objects = pd.read_csv(filepath)
        else:
            self.staff_objects = pd.DataFrame()
        return self.staff_objects
    

    def load_enrollments(self):
        """Load Enrollments (supports 'enrollments.csv' or 'enrollments (3).csv')."""
        # Try common filenames
        paths = [self.data_dir / 'enrollments.csv', self.data_dir / 'enrollments (3).csv']
        for p in paths:
            if p.exists():
                df = pd.read_csv(p, low_memory=False)
                df.columns = df.columns.str.strip()
                # Normalize key columns to numeric where applicable
                if 'enrollment' in df.columns:
                    df['enrollment'] = pd.to_numeric(df['enrollment'], errors='coerce')
                if 'student' in df.columns:
                    df['student'] = pd.to_numeric(df['student'], errors='coerce')
                self.enrollments = df
                return self.enrollments
        # Fallback empty
        self.enrollments = pd.DataFrame()
        return self.enrollments

    def load_invoices(self):
        """Load Invoices."""
        filepath = self.data_dir / 'Invoices.csv'
        if filepath.exists():
            self.invoices = pd.read_csv(filepath)
            self.invoices.columns = self.invoices.columns.str.strip()
        else:
            self.invoices = pd.DataFrame()
        return self.invoices
    
    def load_payments(self):
        """Load Payments."""
        filepath = self.data_dir / 'Payments.csv'
        if filepath.exists():
            self.payments = pd.read_csv(filepath)
            self.payments.columns = self.payments.columns.str.strip()
            
            # Convert date columns
            if 'date' in self.payments.columns:
                self.payments['date'] = pd.to_datetime(self.payments['date'], errors='coerce')
            if 'create date' in self.payments.columns:
                self.payments['create date'] = pd.to_datetime(self.payments['create date'], errors='coerce')
        else:
            self.payments = pd.DataFrame()
        return self.payments
    
    def load_all(self):
        """Load all data files."""
        self.load_students()
        self.load_signoffs()
        self.load_evaluations()
        self.load_events()
        self.load_event_objects()
        self.load_staff_objects()
        self.load_enrollments()
        self.load_invoices()
        self.load_payments()


class StudentProfiler:
    """Creates comprehensive student profiles from loaded data."""
    
    def __init__(self, loader):
        self.loader = loader
    
    def get_student_by_id(self, student_id):
        """Get student record by ID."""
        student = self.loader.students[self.loader.students['id'] == student_id]
        return student.iloc[0] if len(student) > 0 else None
    
    def get_student_by_name(self, name):
        """Get student record by name (partial match)."""
        name_lower = name.lower()
        matches = self.loader.students[
            self.loader.students['full_name'].str.lower().str.contains(name_lower, na=False)
        ]
        return matches
    
    def get_signoffs(self, student_name):
        """Get all signoffs for a student by name."""
        signoffs = self.loader.signoffs[
            self.loader.signoffs['submitter'].str.contains(student_name, case=False, na=False)
        ]
        return signoffs
    
    def get_evaluations(self, student_id):
        """Get all evaluations for a student by ID."""
        evals = self.loader.evaluations[self.loader.evaluations['Student'] == student_id]
        return evals
    
    def get_events_for_student(self, student_id):
        """Get all events associated with a student through evaluations."""
        evals = self.get_evaluations(student_id)
        if len(evals) == 0:
            return pd.DataFrame()
        
        event_ids = evals['event'].unique()
        events = self.loader.events[self.loader.events['event id'].isin(event_ids)]
        return events
    
    def get_event_objects_for_events(self, event_ids):
        """Get all objects/vehicles for given event IDs."""
        if len(event_ids) == 0 or self.loader.event_objects is None or len(self.loader.event_objects) == 0:
            return pd.DataFrame()
        
        # Check if 'event' column exists
        if 'event' not in self.loader.event_objects.columns:
            print(f"⚠️ WARNING: 'event' column not found in event_objects!")
            return pd.DataFrame()
        
        # Match event IDs
        event_ids_numeric = pd.to_numeric(pd.Series(event_ids), errors='coerce').dropna().unique()
        objects = self.loader.event_objects[
            self.loader.event_objects['event'].isin(event_ids_numeric)
        ].copy()
        
        return objects
    
    def get_invoices_for_student(self, student_id):
        """Get all invoices for a student."""
        if self.loader.invoices is None or len(self.loader.invoices) == 0:
            print(f"No invoices loaded")
            return pd.DataFrame()
        
        print(f"Looking for invoices with enrollment = {student_id}")
        print(f"Invoice columns: {self.loader.invoices.columns.tolist()}")
        print(f"Sample enrollments: {self.loader.invoices['enrollment'].head(10).tolist()}")
        print(f"Student ID type: {type(student_id)}")
        print(f"Enrollment column type: {self.loader.invoices['enrollment'].dtype}")
        
        # Link invoices through enrollment (enrollment = student id)
        invoices = self.loader.invoices[
            self.loader.invoices['enrollment'] == student_id
        ].copy()
        
        print(f"Found {len(invoices)} invoices for student {student_id}")
        
        return invoices
    
    def get_payments_for_invoices(self, invoice_ids):
        """Get all payments for given invoice IDs."""
        if len(invoice_ids) == 0 or self.loader.payments is None or len(self.loader.payments) == 0:
            return pd.DataFrame()
        
        payments = self.loader.payments[
            self.loader.payments['invoice'].isin(invoice_ids)
        ].copy()
        
        return payments
    
    def build_complete_profile(self, student_id=None, student_name=None):
    
        """Build a comprehensive profile for a student."""
    
        if student_id:
    
            student_record = self.get_student_by_id(student_id)
    
            if student_record is None:
    
                return None
    
            search_name = f"{student_record['first_name']} {student_record['last_name']}"
    
        elif student_name:
    
            matches = self.get_student_by_name(student_name)
    
            if len(matches) == 0:
    
                return None
    
            elif len(matches) > 1:
    
                return matches
    
            student_record = matches.iloc[0]
    
            student_id = student_record['id']
    
            search_name = student_record['full_name']
    
        else:
    
            return None
    
        
    
        profile = {
    
            'student': student_record,
    
            'signoffs': self.get_signoffs(search_name),
    
            'evaluations': self.get_evaluations(student_id),
    
            'events': self.get_events_for_student(student_id)
    
        }
    
    
    
        # Event objects
    
        if len(profile['events']) > 0:
    
            event_ids = profile['events']['event id'].unique()
    
            profile['event_objects'] = self.get_event_objects_for_events(event_ids)
    
        else:
    
            profile['event_objects'] = pd.DataFrame()
    
    
    
        # --- Correct linkage: enrollments → invoices → payments ---
    
        invoices_df = self.loader.invoices if self.loader.invoices is not None else pd.DataFrame()
    
        payments_df = self.loader.payments if self.loader.payments is not None else pd.DataFrame()
    
        enroll_df = getattr(self.loader, 'enrollments', pd.DataFrame())
    
    
    
        # Normalize
    
        if not enroll_df.empty:
    
            if 'enrollment' in enroll_df.columns:
    
                enroll_df['enrollment'] = pd.to_numeric(enroll_df['enrollment'], errors='coerce')
    
            if 'student' in enroll_df.columns:
    
                enroll_df['student'] = pd.to_numeric(enroll_df['student'], errors='coerce')
    
        if not invoices_df.empty:
    
            if 'enrollment' in invoices_df.columns:
    
                invoices_df['enrollment'] = pd.to_numeric(invoices_df['enrollment'], errors='coerce')
    
            if 'invoice id' in invoices_df.columns:
    
                invoices_df['invoice id'] = pd.to_numeric(invoices_df['invoice id'], errors='coerce')
    
        if not payments_df.empty and 'invoice' in payments_df.columns:
    
            payments_df['invoice'] = pd.to_numeric(payments_df['invoice'], errors='coerce')
    
    
    
        # Default fallbacks
    
        linked_invoices = pd.DataFrame()
    
        linked_payments = pd.DataFrame()
    
    
    
        if not enroll_df.empty and 'student' in enroll_df.columns and 'enrollment' in enroll_df.columns:
    
            student_enrolls = enroll_df[enroll_df['student'] == pd.to_numeric(student_id, errors='coerce')]
    
            if not student_enrolls.empty and not invoices_df.empty and 'enrollment' in invoices_df.columns:
    
                linked_invoices = invoices_df[invoices_df['enrollment'].isin(student_enrolls['enrollment'])].copy()
    
                if not linked_invoices.empty and not payments_df.empty and 'invoice id' in linked_invoices.columns and 'invoice' in payments_df.columns:
    
                    linked_payments = payments_df[payments_df['invoice'].isin(linked_invoices['invoice id'])].copy()
    
    
    
        # If linkage failed, preserve previous behavior
    
        if linked_invoices.empty:
    
            linked_invoices = self.get_invoices_for_student(student_id)
    
            if not linked_invoices.empty and not payments_df.empty and 'invoice id' in linked_invoices.columns and 'invoice' in payments_df.columns:
    
                linked_payments = payments_df[payments_df['invoice'].isin(linked_invoices['invoice id'])].copy()
    
    
    
        profile['invoices'] = linked_invoices if not linked_invoices.empty else pd.DataFrame()
    
        profile['payments'] = linked_payments if not linked_payments.empty else pd.DataFrame()
    
    
    
        return profile

    
    def remove_timezone_from_dataframe(self, df):
        """Remove timezone information from all datetime columns."""
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                if hasattr(df_copy[col].dtype, 'tz') and df_copy[col].dtype.tz is not None:
                    df_copy[col] = df_copy[col].dt.tz_localize(None)
        return df_copy
    
    def build_event_centric_report(self, profile):
        """Build a report organized by events with all related data."""
        if profile is None or isinstance(profile, pd.DataFrame):
            return None
        
        events_data = []
        events = profile['events'].copy()
        
        for idx, event_row in events.iterrows():
            event_id = event_row.get('event id')
            
            # Get trainer name from Staff Objects
            trainer_name = ''
            if self.loader.staff_objects is not None and len(self.loader.staff_objects) > 0:
                staff_for_event = self.loader.staff_objects[
                    (self.loader.staff_objects['event'] == event_id) &
                    (self.loader.staff_objects['state'] == 'active')
                ]
                if len(staff_for_event) > 0:
                    trainer_names = staff_for_event['object'].tolist()
                    trainer_name = ', '.join([str(name) for name in trainer_names if pd.notna(name)])
            
            event_info = {
                'Event ID': event_id,
                'Office': event_row.get('office', ''),
                'Module': event_row.get('module', event_row.get('staff title', '')),
                'Location': event_row.get('location', event_row.get('campus', '')),
                'Event Date/Time': event_row.get('event', ''),
                'Odometer Start': event_row.get('odometer start', ''),
                'Odometer End': event_row.get('odometer end', ''),
                'Progress State': event_row.get('progress state', ''),
                'Staff Sign Off': trainer_name,
                'Campus': event_row.get('campus', ''),
            }
            
            event_evals = profile['evaluations'][
                profile['evaluations']['event'] == event_id
            ].copy()
            
            if len(event_evals) > 0:
                event_info['Evaluation Count'] = len(event_evals)
                event_info['Average Score'] = event_evals['score'].mean()
                
                skills_data = []
                for _, eval_row in event_evals.iterrows():
                    skills_data.append({
                        'Skill Group': eval_row.get('skill group', ''),
                        'Score': eval_row.get('score', ''),
                        'Notes': eval_row.get('notes', eval_row.get('comment', '')),
                    })
                event_info['Skills'] = skills_data
            else:
                event_info['Evaluation Count'] = 0
                event_info['Average Score'] = ''
                event_info['Skills'] = []
            
            # Get event objects/vehicles
            objects_data = []
            
            if len(profile['event_objects']) > 0 and 'event' in profile['event_objects'].columns:
                event_objects_for_this_event = profile['event_objects'][
                    profile['event_objects']['event'] == event_id
                ].copy()
                
                for _, obj_row in event_objects_for_this_event.iterrows():
                    obj_name = obj_row.get('vehicle', '')
                    obj_type = obj_row.get('type', '')
                    
                    if obj_name and str(obj_name).strip():
                        objects_data.append({
                            'Object': str(obj_name).strip(),
                            'Type': str(obj_type).strip() if obj_type else '',
                        })
            
            event_info['Objects'] = objects_data
            event_info['Object Count'] = len(objects_data)
            
            event_start = pd.to_datetime(event_row.get('start'), errors='coerce')
            if pd.notna(event_start):
                event_date = event_start.date()
                signoffs_for_event = profile['signoffs'][
                    pd.to_datetime(profile['signoffs']['date'], errors='coerce').dt.date == event_date
                ].copy()
            else:
                signoffs_for_event = pd.DataFrame()
            
            signoffs_data = []
            for _, sign_row in signoffs_for_event.iterrows():
                signoffs_data.append({
                    'Signature': sign_row.get('signature', ''),
                    'Signoff ID': sign_row.get('signoff id', ''),
                    'State': sign_row.get('state', ''),
                    'Submitter': sign_row.get('submitter', ''),
                })
            event_info['Signoffs'] = signoffs_data
            event_info['Signoff Count'] = len(signoffs_data)
            
            events_data.append(event_info)
        
        return events_data
    
    
    def remove_timezone_from_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ensure Excel-safe datetimes by removing tzinfo."""
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        for col in df.select_dtypes(include=["datetime64[ns, UTC]"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        return df

    def export_event_centric_to_excel(self, profile):
            """Export event-centric report to Excel with professional formatting and return as bytes."""
            if profile is None or isinstance(profile, pd.DataFrame):
                return None
        
            events_data = self.build_event_centric_report(profile)
            if not events_data:
                return None
        
            try:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from openpyxl.formatting.rule import ColorScaleRule
                formatting_available = True
            except ImportError:
                formatting_available = False
        
            output = io.BytesIO()
        
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Ensure timezone-safe data
                for k, v in profile.items():
                    if isinstance(v, pd.DataFrame):
                        profile[k] = self.remove_timezone_from_dataframe(v)
                student = profile['student']
            
                # SHEET 1: Student Profile Cover Page
                info_data = {
                    'Field': ['Student ID', 'Full Name', 'Preferred Name', 'Email', 'Phone', 'Status', 
                              'Code', 'Created Date', '', 
                              'Total Events', 'Total Evaluations', 'Total Signoffs', 'Total Objects Used',
                              'Average Score', 'Agree Signoffs', 'Disagree Signoffs', '',
                              'Total Invoices', 'Total Payments', 'Total Amount Paid'],
                    'Value': [
                        student['id'],
                        f"{student['first_name']} {student['last_name']}",
                        student.get('preferred_name', 'N/A'),
                        student.get('email', 'N/A'),
                        student.get('phone', 'N/A'),
                        student.get('status', 'N/A'),
                        student.get('code', 'N/A'),
                        student.get('created_date', 'N/A'),
                        '',
                        len(profile['events']),
                        len(profile['evaluations']),
                        len(profile['signoffs']),
                        len(profile['event_objects']),
                        f"{profile['evaluations']['score'].mean():.2f}" if len(profile['evaluations']) > 0 else 'N/A',
                        (profile['signoffs']['state'] == 'agree').sum() if len(profile['signoffs']) > 0 else 0,
                        (profile['signoffs']['state'] == 'disagree').sum() if len(profile['signoffs']) > 0 else 0,
                        '',
                        len(profile['invoices']),
                        len(profile['payments']),
                        f"${profile['payments']['payment'].sum():,.2f}" if len(profile['payments']) > 0 else '$0.00',
                    ]
                }
            
                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name='👤 Student Profile', index=False)
            
                if formatting_available:
                    ws_info = writer.sheets['👤 Student Profile']
                    ws_info.insert_rows(1)
                    ws_info.merge_cells('A1:B1')
                    ws_info['A1'] = f"STUDENT PROFILE: {student['first_name']} {student['last_name']}"
                    ws_info['A1'].font = Font(size=16, bold=True, color='FFFFFF')
                    ws_info['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    ws_info['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_info.row_dimensions[1].height = 35
                
                    ws_info['A2'].font = Font(bold=True, size=11, color='FFFFFF')
                    ws_info['A2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    ws_info['B2'].font = Font(bold=True, size=11, color='FFFFFF')
                    ws_info['B2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                
                    info_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    stats_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                
                    for row_idx in range(3, ws_info.max_row + 1):
                        field_cell = ws_info[f'A{row_idx}']
                        value_cell = ws_info[f'B{row_idx}']
                        field_cell.font = Font(bold=True, size=10)
                        field_cell.alignment = Alignment(horizontal='right', vertical='center')
                        value_cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                        if row_idx >= 12:
                            field_cell.fill = stats_fill
                            value_cell.fill = stats_fill
                        else:
                            field_cell.fill = info_fill
                            value_cell.fill = info_fill
                
                    ws_info.column_dimensions['A'].width = 25
                    ws_info.column_dimensions['B'].width = 40

                    # ========== NEW SECTION: INVOICES AND PAYMENTS ==========
                    from openpyxl.utils import get_column_letter
                    current_row = ws_info.max_row + 2
                
                    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    cell_font = Font(size=10)
                    alt_fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
                
                    # ----- INVOICES SECTION -----
                    ws_info.merge_cells(f'A{current_row}:B{current_row}')
                    ws_info[f'A{current_row}'] = "🧾 INVOICES"
                    ws_info[f'A{current_row}'].font = Font(bold=True, size=12, color='FFFFFF')
                    ws_info[f'A{current_row}'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                    ws_info[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_info.row_dimensions[current_row].height = 22
                    current_row += 1
                
                    invoices = profile.get('invoices', pd.DataFrame())
                    if not invoices.empty:
                        cols = invoices.columns.tolist()
                        header_row = current_row
                        for col_idx, col_name in enumerate(cols, start=1):
                            cell = ws_info.cell(row=header_row, column=col_idx)
                            cell.value = col_name
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                
                        for _, row in invoices.iterrows():
                            for col_idx, col_name in enumerate(cols, start=1):
                                value = row.get(col_name, "")
                                cell = ws_info.cell(row=current_row, column=col_idx)
                                cell.value = value
                                cell.font = cell_font
                                cell.fill = alt_fill if current_row % 2 == 0 else PatternFill()
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            current_row += 1
                    else:
                        ws_info[f'A{current_row}'] = "No records found"
                        ws_info[f'A{current_row}'].font = Font(italic=True, color='808080')
                        current_row += 2
                
                    # ----- PAYMENTS SECTION -----
                    ws_info.merge_cells(f'A{current_row}:B{current_row}')
                    ws_info[f'A{current_row}'] = "💳 PAYMENTS"
                    ws_info[f'A{current_row}'].font = Font(bold=True, size=12, color='FFFFFF')
                    ws_info[f'A{current_row}'].fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                    ws_info[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_info.row_dimensions[current_row].height = 22
                    current_row += 1
                
                    payments = profile.get('payments', pd.DataFrame())
                    if not payments.empty:
                        cols = payments.columns.tolist()
                        header_row = current_row
                        for col_idx, col_name in enumerate(cols, start=1):
                            cell = ws_info.cell(row=header_row, column=col_idx)
                            cell.value = col_name
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                
                        for _, row in payments.iterrows():
                            for col_idx, col_name in enumerate(cols, start=1):
                                value = row.get(col_name, "")
                                cell = ws_info.cell(row=current_row, column=col_idx)
                                cell.value = value
                                cell.font = cell_font
                                cell.fill = alt_fill if current_row % 2 == 0 else PatternFill()
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            current_row += 1
                    else:
                        ws_info[f'A{current_row}'] = "No records found"
                        ws_info[f'A{current_row}'].font = Font(italic=True, color='808080')
                        current_row += 1
                
                    # Auto-width adjustment
                    for col in ws_info.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value)) > max_len:
                                    max_len = len(str(cell.value))
                            except:
                                pass
                        ws_info.column_dimensions[col_letter].width = min(max_len + 2, 40)
                    # ========== END NEW SECTION ==========

            
                # SHEET 2: Events Summary
                summary_rows = []
                for event_info in events_data:
                    summary_rows.append({
                        'Event ID': event_info['Event ID'],
                        'Office': event_info['Office'],
                        'Module': event_info['Module'],
                        'Location': event_info['Location'],
                        'Event Date/Time': event_info['Event Date/Time'],
                        'Progress State': event_info['Progress State'],
                        'Evaluation Count': event_info['Evaluation Count'],
                        'Average Score': event_info['Average Score'],
                        'Object Count': event_info['Object Count'],
                        'Signoff Count': event_info['Signoff Count'],
                    })
            
                summary_df = pd.DataFrame(summary_rows)
                summary_df = self.remove_timezone_from_dataframe(summary_df)
            
                for col in summary_df.columns:
                    if summary_df[col].dtype == 'object':
                        summary_df[col] = summary_df[col].apply(
                            lambda x: str(x) if pd.notna(x) and hasattr(x, 'tzinfo') else x
                        )
            
                summary_df.to_excel(writer, sheet_name='📊 Events Summary', index=False)
            
                if formatting_available:
                    ws_summary = writer.sheets['📊 Events Summary']
                    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True, size=11)
                
                    for cell in ws_summary[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                    for column in ws_summary.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_summary.column_dimensions[column_letter].width = adjusted_width
                
                    ws_summary.freeze_panes = 'A2'
                    ws_summary.auto_filter.ref = ws_summary.dimensions
                
                    score_col = None
                    for idx, cell in enumerate(ws_summary[1], 1):
                        if cell.value == 'Average Score':
                            score_col = get_column_letter(idx)
                            break
                
                    if score_col:
                        ws_summary.conditional_formatting.add(
                            f'{score_col}2:{score_col}{len(summary_df)+1}',
                            ColorScaleRule(
                                start_type='num', start_value=0, start_color='F8696B',
                                mid_type='num', mid_value=3, mid_color='FFEB84',
                                end_type='num', end_value=5, end_color='63BE7B'
                            )
                        )
            
                # SHEET 3: Detailed Report (Compact View with Color Coding)
                compact_rows = []
                for event_info in events_data:
                    # First row has all event info
                    first_row = {
                        'Event ID': event_info['Event ID'],
                        'Office': event_info['Office'],
                        'Module': event_info['Module'],
                        'Location': event_info['Location'],
                        'Event Date/Time': event_info['Event Date/Time'],
                        'Odometer Start': event_info['Odometer Start'],
                        'Odometer End': event_info['Odometer End'],
                        'Progress State': event_info['Progress State'],
                        'Staff Sign Off': event_info['Staff Sign Off'],
                        'Skill Group': '',
                        'Score': '',
                        'Notes': '',
                        'Object': '',
                        'Object Type': '',
                        'Signature': '',
                        'Signoff ID': '',
                        'Signoff State': '',
                        'Submitter': '',
                    }
                
                    # Add first object/skill/signoff to first row if they exist
                    if len(event_info['Skills']) > 0:
                        first_row['Skill Group'] = event_info['Skills'][0]['Skill Group']
                        first_row['Score'] = event_info['Skills'][0]['Score']
                        first_row['Notes'] = event_info['Skills'][0]['Notes']
                
                    if len(event_info['Objects']) > 0:
                        first_row['Object'] = event_info['Objects'][0]['Object']
                        first_row['Object Type'] = event_info['Objects'][0]['Type']
                
                    if len(event_info['Signoffs']) > 0:
                        first_row['Signature'] = event_info['Signoffs'][0]['Signature']
                        first_row['Signoff ID'] = event_info['Signoffs'][0]['Signoff ID']
                        first_row['Signoff State'] = event_info['Signoffs'][0]['State']
                        first_row['Submitter'] = event_info['Signoffs'][0]['Submitter']
                
                    compact_rows.append(first_row)
                
                    # Calculate how many additional rows we need (start from index 1 since 0 is in first row)
                    max_items = max(
                        len(event_info['Skills']),
                        len(event_info['Objects']),
                        len(event_info['Signoffs'])
                    )
                
                    # Add additional rows for remaining items
                    for i in range(1, max_items):
                        row = {
                            'Event ID': '',
                            'Office': '',
                            'Module': '',
                            'Location': '',
                            'Event Date/Time': '',
                            'Odometer Start': '',
                            'Odometer End': '',
                            'Progress State': '',
                            'Staff Sign Off': '',
                            'Skill Group': '',
                            'Score': '',
                            'Notes': '',
                            'Object': '',
                            'Object Type': '',
                            'Signature': '',
                            'Signoff ID': '',
                            'Signoff State': '',
                            'Submitter': '',
                        }
                    
                        if i < len(event_info['Skills']):
                            row['Skill Group'] = event_info['Skills'][i]['Skill Group']
                            row['Score'] = event_info['Skills'][i]['Score']
                            row['Notes'] = event_info['Skills'][i]['Notes']
                    
                        if i < len(event_info['Objects']):
                            row['Object'] = event_info['Objects'][i]['Object']
                            row['Object Type'] = event_info['Objects'][i]['Type']
                    
                        if i < len(event_info['Signoffs']):
                            row['Signature'] = event_info['Signoffs'][i]['Signature']
                            row['Signoff ID'] = event_info['Signoffs'][i]['Signoff ID']
                            row['Signoff State'] = event_info['Signoffs'][i]['State']
                            row['Submitter'] = event_info['Signoffs'][i]['Submitter']
                    
                        compact_rows.append(row)
                
                    # Add blank row between events
                    compact_rows.append({})
            
                compact_df = pd.DataFrame(compact_rows)
                compact_df = self.remove_timezone_from_dataframe(compact_df)
            
                for col in compact_df.columns:
                    if compact_df[col].dtype == 'object':
                        compact_df[col] = compact_df[col].apply(
                            lambda x: str(x) if pd.notna(x) and hasattr(x, 'tzinfo') else x
                        )
            
                compact_df.to_excel(writer, sheet_name='📋 Detailed Report', index=False)
            
                if formatting_available:
                    ws_compact = writer.sheets['📋 Detailed Report']
                
                    event_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    eval_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    object_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    signoff_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                
                    for cell in ws_compact[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                    thin_border = Border(
                        left=Side(style='thin', color='D3D3D3'),
                        right=Side(style='thin', color='D3D3D3'),
                        top=Side(style='thin', color='D3D3D3'),
                        bottom=Side(style='thin', color='D3D3D3')
                    )
                
                    event_font = Font(bold=True, size=10)
                
                    for row_idx, row in enumerate(ws_compact.iter_rows(min_row=2, max_row=ws_compact.max_row), start=2):
                        if row[0].value and str(row[0].value).strip():
                            for cell in row[:9]:
                                cell.fill = event_fill
                                cell.font = event_font
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical='center')
                        else:
                            for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical='center')
                        
                            if row[9].value:
                                for cell in row[9:12]:
                                    cell.fill = eval_fill
                        
                            if row[12].value:
                                for cell in row[12:14]:
                                    cell.fill = object_fill
                        
                            if row[14].value:
                                for cell in row[14:18]:
                                    cell.fill = signoff_fill
                
                    for column in ws_compact.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        ws_compact.column_dimensions[column_letter].width = adjusted_width
                
                    ws_compact.freeze_panes = 'A2'
                    ws_compact.auto_filter.ref = ws_compact.dimensions
                    ws_compact.row_dimensions[1].height = 30
            
                # SHEET 4: Clean Table View (Option 5)
                clean_rows = []
                for event_info in events_data:
                    # Calculate hours if odometer data exists
                    hours_worked = ''
                    odo_start = event_info.get('Odometer Start', '')
                    odo_end = event_info.get('Odometer End', '')
                    if odo_start and odo_end:
                        try:
                            hours = float(odo_end) - float(odo_start)
                            hours_worked = f"{hours:.1f}"
                        except:
                            hours_worked = ''
                
                    # Get equipment name
                    equipment = ''
                    if len(event_info['Objects']) > 0:
                        equipment = event_info['Objects'][0]['Object']
                
                    clean_rows.append({
                        'Date': event_info['Event Date/Time'],
                        'Module': event_info['Module'],
                        'Location': event_info['Location'],
                        'Equipment': equipment,
                        'Hours': hours_worked,
                        'Odometer': f"{odo_start} → {odo_end}" if odo_start and odo_end else '',
                        'Avg Score': f"{event_info['Average Score']:.1f}" if event_info['Average Score'] else '',
                        'Skills Evaluated': event_info['Evaluation Count'],
                        'Instructor': event_info['Staff Sign Off'],
                        'Status': '✓ Complete' if event_info['Progress State'] == 'finished' else event_info['Progress State'],
                    })
            
                clean_df = pd.DataFrame(clean_rows)
                clean_df = self.remove_timezone_from_dataframe(clean_df)
                clean_df.to_excel(writer, sheet_name='📋 Quick View', index=False)
            
                if formatting_available:
                    ws_clean = writer.sheets['📋 Quick View']
                
                    # Header formatting
                    header_fill = PatternFill(start_color='2C5F2D', end_color='2C5F2D', fill_type='solid')
                    header_font = Font(color='FFFFFF', bold=True, size=11)
                
                    for cell in ws_clean[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                    # Alternating row colors for readability
                    light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                    for row_idx in range(2, ws_clean.max_row + 1):
                        fill = light_fill if row_idx % 2 == 0 else white_fill
                        for cell in ws_clean[row_idx]:
                            cell.fill = fill
                            cell.alignment = Alignment(vertical='center')
                            cell.border = Border(
                                bottom=Side(style='thin', color='E0E0E0')
                            )
                
                    # Column widths
                    ws_clean.column_dimensions['A'].width = 20
                    ws_clean.column_dimensions['B'].width = 25
                    ws_clean.column_dimensions['C'].width = 20
                    ws_clean.column_dimensions['D'].width = 15
                    ws_clean.column_dimensions['E'].width = 10
                    ws_clean.column_dimensions['F'].width = 20
                    ws_clean.column_dimensions['G'].width = 12
                    ws_clean.column_dimensions['H'].width = 15
                    ws_clean.column_dimensions['I'].width = 20
                    ws_clean.column_dimensions['J'].width = 15
                
                    ws_clean.freeze_panes = 'A2'
                    ws_clean.auto_filter.ref = ws_clean.dimensions
            
                
            # SHEET 5: Card View (Multi-sheet by Module Category with 📄 prefix)
            events_df = profile.get('events', pd.DataFrame()).copy()
            if not events_df.empty and 'module' in events_df.columns:
                import re
                # Clean and extract main category (case-insensitive, ignore punctuation)
                events_df['module_clean'] = events_df['module'].astype(str).str.strip()
                events_df['main_category'] = events_df['module_clean'].str.extract(r'^([A-Za-z]+)')[0].str.upper()

                # Extract numeric part for sorting
                def extract_number(s):
                    match = re.search(r'(\d+)', str(s))
                    return int(match.group(1)) if match else 9999

                events_df['module_num'] = events_df['module_clean'].apply(extract_number)

                grouped = events_df.groupby('main_category', dropna=False)
                for cat, g in grouped:
                    category = str(cat).strip().upper() if pd.notna(cat) and str(cat).strip() else 'OTHER'
                    g_sorted = g.sort_values('module_num', ascending=True).copy()
                    sheet_name = f"📄 {category}"[:31]

                    g_sorted.to_excel(writer, sheet_name=sheet_name, index=False)

                    if formatting_available:
                        ws = writer.sheets[sheet_name]
                        from openpyxl.styles import PatternFill, Font, Alignment

                        blue_dark = '1F4E78'
                        blue_med = '305496'
                        gray_light = 'D9D9D9'

                        for row_idx in range(2, ws.max_row + 1):
                            module_val = str(ws.cell(row_idx, 2).value or '').strip()
                            if module_val:
                                # Row 1: Module header
                                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                                c = ws.cell(row=row_idx, column=1)
                                c.value = module_val
                                c.font = Font(bold=True, color='FFFFFF')
                                c.fill = PatternFill(start_color=blue_dark, end_color=blue_dark, fill_type='solid')
                                c.alignment = Alignment(horizontal='center', vertical='center')
                                row_idx += 1

                                # Row 2: Date/Time header
                                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                                c = ws.cell(row=row_idx, column=1)
                                c.font = Font(bold=True, color='FFFFFF')
                                c.fill = PatternFill(start_color=blue_med, end_color=blue_med, fill_type='solid')
                                c.alignment = Alignment(horizontal='center', vertical='center')
                                row_idx += 1

                                # Row 3: Event ID
                                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                                c = ws.cell(row=row_idx, column=1)
                                c.font = Font(size=10, color='808080')
                                c.fill = PatternFill(start_color=gray_light, end_color=gray_light, fill_type='solid')
                                c.alignment = Alignment(horizontal='center', vertical='center')
